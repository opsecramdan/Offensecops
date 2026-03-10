from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_
from sqlalchemy.orm import selectinload
from pydantic import BaseModel
import csv, io, socket, json

from app.db.session import get_db
from app.db.models import Target, TargetGroup, User
from app.api.deps import get_current_user

router = APIRouter()


# ── Schemas ───────────────────────────────────────────────────
class GroupCreate(BaseModel):
    name: str
    description: Optional[str] = None
    color: str = "#6366f1"

class GroupUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    color: Optional[str] = None

class TargetCreate(BaseModel):
    value: str
    type: str
    ip_address: Optional[str] = None
    group_id: Optional[str] = None
    owner: Optional[str] = None
    criticality: str = "medium"
    tags: List[str] = []
    notes: Optional[str] = None
    scope_status: str = "in_scope"

class TargetBulkCreate(BaseModel):
    targets: List[dict]
    group_id: Optional[str] = None
    criticality: str = "medium"
    scope_status: str = "in_scope"
    owner: Optional[str] = None

class TargetUpdate(BaseModel):
    ip_address: Optional[str] = None
    group_id: Optional[str] = None
    owner: Optional[str] = None
    criticality: Optional[str] = None
    tags: Optional[List[str]] = None
    notes: Optional[str] = None
    scope_status: Optional[str] = None


# ── Serializers ───────────────────────────────────────────────
def group_to_dict(g: TargetGroup) -> dict:
    return {
        "id": g.id, "name": g.name, "description": g.description,
        "color": g.color or "#6366f1",
        "created_at": g.created_at.isoformat() if g.created_at else None,
    }

def target_to_dict(t: Target) -> dict:
    return {
        "id": t.id, "value": t.value, "type": t.type,
        "ip_address": t.ip_address,
        "group_id": t.group_id,
        "group": group_to_dict(t.group) if t.group else None,
        "owner": t.owner, "criticality": t.criticality,
        "tags": t.tags or [], "notes": t.notes,
        "scope_status": t.scope_status, "created_by": t.created_by,
        "created_at": t.created_at.isoformat() if t.created_at else None,
        "updated_at": t.updated_at.isoformat() if t.updated_at else None,
    }


# ══════════════════════════════════════════════════════════════
# GROUP ROUTES (before /{target_id} to avoid conflicts)
# ══════════════════════════════════════════════════════════════

@router.get("/groups")
async def list_groups(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(TargetGroup).order_by(TargetGroup.name))
    groups = result.scalars().all()
    # Include target count per group
    out = []
    for g in groups:
        count_q = select(func.count()).where(Target.group_id == g.id)
        count = (await db.execute(count_q)).scalar()
        d = group_to_dict(g)
        d["target_count"] = count
        out.append(d)
    return out


@router.post("/groups", status_code=201)
async def create_group(
    data: GroupCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    existing = await db.execute(select(TargetGroup).where(TargetGroup.name == data.name))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail=f"Group '{data.name}' already exists")
    group = TargetGroup(
        name=data.name, description=data.description,
        color=data.color, created_by=current_user.id,
    )
    db.add(group)
    await db.commit()
    await db.refresh(group)
    return group_to_dict(group)


@router.patch("/groups/{group_id}")
async def update_group(
    group_id: str, data: GroupUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(TargetGroup).where(TargetGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    for field, val in data.model_dump(exclude_none=True).items():
        setattr(group, field, val)
    await db.commit()
    await db.refresh(group)
    return group_to_dict(group)


@router.delete("/groups/{group_id}", status_code=204)
async def delete_group(
    group_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(TargetGroup).where(TargetGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    # Unassign targets from this group
    targets = await db.execute(select(Target).where(Target.group_id == group_id))
    for t in targets.scalars().all():
        t.group_id = None
    await db.delete(group)
    await db.commit()


# ══════════════════════════════════════════════════════════════
# SPECIFIC ROUTES (before /{target_id})
# ══════════════════════════════════════════════════════════════

@router.get("/resolve/{hostname}")
async def resolve_hostname(
    hostname: str,
    current_user: User = Depends(get_current_user),
):
    import asyncio
    clean = hostname.replace("http://","").replace("https://","").split("/")[0].split(":")[0]
    try:
        loop = asyncio.get_event_loop()
        ip = await loop.run_in_executor(None, socket.gethostbyname, clean)
        return {"hostname": clean, "ip": ip, "resolved": True}
    except Exception:
        return {"hostname": clean, "ip": None, "resolved": False}


@router.post("/bulk", status_code=201)
async def bulk_create_targets(
    data: TargetBulkCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    created, skipped, items = 0, 0, []
    for t in data.targets:
        value = str(t.get("value", "")).strip()
        if not value:
            continue
        existing = await db.execute(select(Target).where(Target.value == value))
        if existing.scalar_one_or_none():
            skipped += 1
            continue
        ip = t.get("ip_address") or t.get("ip")
        if not ip and t.get("type") in ("domain", "subdomain"):
            try:
                ip = socket.gethostbyname(value)
            except Exception:
                ip = None
        tags = list(t.get("tags", []))
        source = t.get("source", "")
        if source and source not in tags:
            tags.append(source)
        target = Target(
            value=value, type=t.get("type", "subdomain"), ip_address=ip,
            group_id=data.group_id, owner=data.owner,
            criticality=data.criticality, tags=tags,
            scope_status=data.scope_status, created_by=current_user.id,
        )
        db.add(target)
        items.append(value)
        created += 1
    await db.commit()
    return {"created": created, "skipped": skipped, "items": items}


@router.get("/export")
async def export_targets(
    format: str = Query("csv", regex="^(csv|txt|xlsx)$"),
    group_id: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    criticality: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export targets as CSV / TXT / Excel — respects active filters"""
    query = select(Target)
    if group_id:
        query = query.where(Target.group_id == group_id)
    if type:
        query = query.where(Target.type == type)
    if criticality:
        query = query.where(Target.criticality == criticality)
    if search:
        query = query.where(Target.value.ilike(f"%{search}%"))
    query = query.order_by(Target.value)
    result = await db.execute(query)
    targets = result.scalars().all()

    if format == "txt":
        lines = "\n".join(t.value for t in targets)
        return StreamingResponse(
            io.BytesIO(lines.encode()),
            media_type="text/plain",
            headers={"Content-Disposition": "attachment; filename=targets.txt"},
        )

    elif format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["value", "type", "ip_address", "group", "criticality", "scope_status", "tags", "owner", "notes"])
        for t in targets:
            writer.writerow([
                t.value, t.type, t.ip_address or "",
                t.group.name if t.group else "",
                t.criticality, t.scope_status,
                ",".join(t.tags or []), t.owner or "", t.notes or "",
            ])
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode()),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=targets.csv"},
        )

    elif format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Targets"

        headers = ["Value", "Type", "IP Address", "Group", "Criticality", "Scope", "Tags", "Owner", "Notes"]
        header_fill = PatternFill(start_color="1e1e2e", end_color="1e1e2e", fill_type="solid")
        header_font = Font(bold=True, color="89b4fa")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        crit_colors = {
            "critical": "ff5f5f", "high": "ff9f43",
            "medium": "ffd43b", "low": "a9e34b", "informational": "74c7ec",
        }
        for row, t in enumerate(targets, 2):
            ws.cell(row=row, column=1, value=t.value)
            ws.cell(row=row, column=2, value=t.type)
            ws.cell(row=row, column=3, value=t.ip_address or "")
            ws.cell(row=row, column=4, value=t.group.name if t.group else "")
            crit_cell = ws.cell(row=row, column=5, value=t.criticality)
            color = crit_colors.get(t.criticality, "ffffff")
            crit_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=row, column=6, value=t.scope_status)
            ws.cell(row=row, column=7, value=",".join(t.tags or []))
            ws.cell(row=row, column=8, value=t.owner or "")
            ws.cell(row=row, column=9, value=t.notes or "")

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=targets.xlsx"},
        )


@router.post("/import/csv", status_code=201)
async def import_targets_csv(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    content = await file.read()
    reader = csv.DictReader(io.StringIO(content.decode("utf-8")))
    created, skipped, errors = 0, 0, []
    for i, row in enumerate(reader):
        try:
            value = row.get("value", "").strip()
            if not value:
                continue
            existing = await db.execute(select(Target).where(Target.value == value))
            if existing.scalar_one_or_none():
                skipped += 1
                continue
            tags_raw = row.get("tags", "")
            tags = [t.strip() for t in tags_raw.split(",") if t.strip()]
            db.add(Target(
                value=value, type=row.get("type", "domain"), owner=row.get("owner"),
                criticality=row.get("criticality", "medium"), tags=tags,
                scope_status=row.get("scope_status", "in_scope"), created_by=current_user.id,
            ))
            created += 1
        except Exception as e:
            errors.append(f"Row {i+2}: {str(e)}")
    await db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


# ══════════════════════════════════════════════════════════════
# COLLECTION ROUTES
# ══════════════════════════════════════════════════════════════

@router.get("/")
async def list_targets(
    search: Optional[str] = Query(None),
    criticality: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    scope_status: Optional[str] = Query(None),
    group_id: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = select(Target).options(selectinload(Target.group))
    if search:
        query = query.where(or_(Target.value.ilike(f"%{search}%"), Target.owner.ilike(f"%{search}%")))
    if criticality:
        query = query.where(Target.criticality == criticality)
    if type:
        query = query.where(Target.type == type)
    if scope_status:
        query = query.where(Target.scope_status == scope_status)
    if group_id:
        query = query.where(Target.group_id == group_id)
    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar()
    query = query.order_by(Target.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(query)
    targets = result.scalars().all()
    return {"total": total, "items": [target_to_dict(t) for t in targets], "skip": skip, "limit": limit}


@router.post("/", status_code=201)
async def create_target(
    data: TargetCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    existing = await db.execute(select(Target).where(Target.value == data.value))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=409, detail=f"Target '{data.value}' already exists")
    target = Target(
        value=data.value.strip(), type=data.type, ip_address=data.ip_address or None,
        group_id=data.group_id or None, owner=data.owner or None,
        criticality=data.criticality, tags=data.tags,
        notes=data.notes or None, scope_status=data.scope_status,
        created_by=current_user.id,
    )
    db.add(target)
    await db.commit()
    await db.refresh(target)
    # Reload with relationship
    result = await db.execute(select(Target).options(selectinload(Target.group)).where(Target.id == target.id))
    target = result.scalar_one()
    return target_to_dict(target)


# ══════════════════════════════════════════════════════════════
# ITEM ROUTES — /{target_id} MUST BE LAST
# ══════════════════════════════════════════════════════════════

@router.get("/{target_id}")
async def get_target(
    target_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Target).options(selectinload(Target.group)).where(Target.id == target_id))
    target = result.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="Target not found")
    return target_to_dict(target)


@router.patch("/{target_id}")
async def update_target(
    target_id: str, data: TargetUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Target).where(Target.id == target_id))
    target = result.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="Target not found")
    for field, val in data.model_dump(exclude_none=True).items():
        setattr(target, field, val)
    await db.commit()
    result = await db.execute(select(Target).options(selectinload(Target.group)).where(Target.id == target_id))
    target = result.scalar_one()
    return target_to_dict(target)


@router.delete("/{target_id}", status_code=204)
async def delete_target(
    target_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Target).where(Target.id == target_id))
    target = result.scalar_one_or_none()
    if not target:
        raise HTTPException(status_code=404, detail="Target not found")
    await db.delete(target)
    await db.commit()
