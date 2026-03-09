from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete
from pydantic import BaseModel
from datetime import datetime, timezone
import uuid, json, io

from app.db.session import get_db
from app.db.models import VulnCompany, VulnReport, VulnStatus
from app.api.deps import get_current_user
from app.db.models import User

router = APIRouter()

# ── Schemas ───────────────────────────────────────────────────
class CompanyCreate(BaseModel):
    name: str
    code: str
    color: str = "#6366f1"
    description: str = ""

class StatusCreate(BaseModel):
    name: str
    color: str = "#74c7ec"

class VulnReportCreate(BaseModel):
    company_id: str
    no: Optional[int] = None
    vuln_code: Optional[str] = None
    vuln_id: Optional[str] = None
    vuln_members: Optional[str] = None
    vuln_name: str
    description: Optional[str] = None
    severity: str = "medium"
    cvss_vector: Optional[str] = None
    cvss_score: Optional[float] = None
    cvss_version: str = "3.1"
    impact: Optional[str] = None
    mitigation: Optional[str] = None
    status: str = "Open"
    finding_date: Optional[str] = None
    resolution_date: Optional[str] = None
    fixing_date: Optional[str] = None
    referensi: Optional[str] = None
    note: Optional[str] = None
    product_id: Optional[str] = None

class VulnReportUpdate(BaseModel):
    no: Optional[int] = None
    vuln_code: Optional[str] = None
    vuln_id: Optional[str] = None
    vuln_members: Optional[str] = None
    vuln_name: Optional[str] = None
    description: Optional[str] = None
    severity: Optional[str] = None
    cvss_vector: Optional[str] = None
    cvss_score: Optional[float] = None
    cvss_version: Optional[str] = None
    impact: Optional[str] = None
    mitigation: Optional[str] = None
    status: Optional[str] = None
    finding_date: Optional[str] = None
    resolution_date: Optional[str] = None
    fixing_date: Optional[str] = None
    referensi: Optional[str] = None
    note: Optional[str] = None
    product_id: Optional[str] = None

def parse_date(s) -> Optional[datetime]:
    if not s: return None
    # Handle Excel serial date number (days since 1900-01-01)
    try:
        n = float(str(s).strip())
        if 1000 < n < 100000:  # reasonable Excel date range
            from datetime import timedelta
            base = datetime(1899, 12, 30, tzinfo=timezone.utc)
            return base + timedelta(days=n)
    except: pass
    # Handle string dates
    s = str(s).strip()
    if not s or s.lower() in ('none','null','-'): return None
    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S", "%d-%m-%Y", "%Y/%m/%d"]:
        try: return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except: pass
    return None

# ── Companies ─────────────────────────────────────────────────
@router.get("/companies")
async def get_companies(db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(VulnCompany).order_by(VulnCompany.created_at))
    companies = result.scalars().all()
    out = []
    for c in companies:
        # Count reports
        cnt = await db.execute(select(func.count(VulnReport.id)).where(VulnReport.company_id == c.id))
        total = cnt.scalar() or 0
        # Status breakdown
        status_q = await db.execute(
            select(VulnReport.status, func.count(VulnReport.id))
            .where(VulnReport.company_id == c.id)
            .group_by(VulnReport.status)
        )
        statuses = {row[0]: row[1] for row in status_q}
        out.append({
            "id": c.id, "name": c.name, "code": c.code,
            "color": c.color, "description": c.description,
            "total": total, "statuses": statuses,
            "created_at": c.created_at.isoformat() if c.created_at else None,
        })
    return out

@router.post("/companies")
async def create_company(data: CompanyCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    company = VulnCompany(
        id=str(uuid.uuid4()),
        name=data.name,
        code=data.code.upper(),
        color=data.color,
        description=data.description,
        created_at=datetime.now(timezone.utc),
    )
    db.add(company)
    await db.commit()
    await db.refresh(company)
    return {"id": company.id, "name": company.name, "code": company.code, "color": company.color}

@router.delete("/companies/{company_id}")
async def delete_company(company_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    await db.execute(delete(VulnReport).where(VulnReport.company_id == company_id))
    await db.execute(delete(VulnCompany).where(VulnCompany.id == company_id))
    await db.commit()
    return {"ok": True}

# ── Statuses ──────────────────────────────────────────────────
@router.get("/statuses")
async def get_statuses(db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(VulnStatus).order_by(VulnStatus.created_at))
    return [{"id": s.id, "name": s.name, "color": s.color, "is_default": s.is_default} for s in result.scalars()]

@router.post("/statuses")
async def create_status(data: StatusCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    s = VulnStatus(id=str(uuid.uuid4()), name=data.name, color=data.color, created_at=datetime.now(timezone.utc))
    db.add(s)
    await db.commit()
    return {"id": s.id, "name": s.name, "color": s.color}

@router.delete("/statuses/{status_id}")
async def delete_status(status_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    await db.execute(delete(VulnStatus).where(VulnStatus.id == status_id))
    await db.commit()
    return {"ok": True}

# ── Reports ───────────────────────────────────────────────────
@router.get("/reports")
async def get_reports(
    company_id: Optional[str] = None,
    severity: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    product_id: Optional[str] = None,
    year: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    q = select(VulnReport).order_by(VulnReport.no, VulnReport.created_at)
    if company_id:  q = q.where(VulnReport.company_id == company_id)
    if severity:    q = q.where(VulnReport.severity == severity)
    if status:      q = q.where(VulnReport.status == status)
    if product_id:  q = q.where(VulnReport.product_id == product_id)
    if year:
        try:
            y = int(year)
            q = q.where(VulnReport.no == y)
        except: pass
    if search:
        q = q.where(
            VulnReport.vuln_name.ilike(f"%{search}%") |
            VulnReport.vuln_id.ilike(f"%{search}%") |
            VulnReport.vuln_code.ilike(f"%{search}%") |
            VulnReport.description.ilike(f"%{search}%")
        )
    result = await db.execute(q)
    reports = result.scalars().all()
    return [_report_dict(r) for r in reports]

@router.post("/reports")
async def create_report(data: VulnReportCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    r = VulnReport(
        id=str(uuid.uuid4()),
        company_id=data.company_id,
        no=data.no,
        vuln_code=data.vuln_code,
        vuln_id=data.vuln_id,
        vuln_members=data.vuln_members,
        vuln_name=data.vuln_name,
        description=data.description,
        severity=data.severity,
        cvss_vector=data.cvss_vector,
        cvss_score=data.cvss_score,
        cvss_version=data.cvss_version,
        impact=data.impact,
        mitigation=data.mitigation,
        status=data.status,
        finding_date=parse_date(data.finding_date),
        resolution_date=parse_date(data.resolution_date),
        fixing_date=parse_date(data.fixing_date),
        referensi=data.referensi,
        note=data.note,
        product_id=data.product_id,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
    )
    db.add(r)
    await db.commit()
    await db.refresh(r)
    return _report_dict(r)

@router.put("/reports/{report_id}")
async def update_report(report_id: str, data: VulnReportUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(VulnReport).where(VulnReport.id == report_id))
    r = result.scalar_one_or_none()
    if not r: raise HTTPException(404, "Not found")
    for field, val in data.model_dump(exclude_none=True).items():
        if field in ("finding_date", "resolution_date", "fixing_date"):
            setattr(r, field, parse_date(val))
        else:
            setattr(r, field, val)
    r.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(r)
    return _report_dict(r)

@router.delete("/reports/{report_id}")
async def delete_report(report_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    await db.execute(delete(VulnReport).where(VulnReport.id == report_id))
    await db.commit()
    return {"ok": True}

# ── Import Excel ──────────────────────────────────────────────
@router.post("/reports/import")
async def import_excel(
    company_id: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    print("Headers detected:", headers)

    def get_col(row, name_variants):
        for v in name_variants:
            if v in headers:
                idx = headers.index(v)
                cell = row[idx]
                return str(cell.value).strip() if cell.value is not None else None
        return None

    imported = 0
    for row in ws.iter_rows(min_row=2):
        if all(c.value is None for c in row): continue
        name = get_col(row, ["vulnerability name","vuln name","vuln_name","name","title"])
        if not name: continue
        r = VulnReport(
            id=str(uuid.uuid4()),
            company_id=company_id,
            no=_to_int(get_col(row, ["no","number","#","periode","period","year","tahun"])),
            vuln_code=get_col(row, ["vulnerability code","vuln code","vuln_code","code","no.vulnerability code","vulnerability_code"]),
            vuln_id=get_col(row, ["vulnerability id","vuln id","vuln_id","vulnerability_id"]),
            vuln_members=get_col(row, ["vulnerability members","members","vuln members","vulnerabilty members","member"]),
            vuln_name=name,
            description=get_col(row, ["description","deskripsi","vulnerability description","vuln description"]),
            severity=_norm_sev(get_col(row, ["severity","tingkat"])),
            cvss_vector=get_col(row, ["cvss vector","cvss 3.1","cvss_vector","vector"]),
            cvss_score=_to_float(get_col(row, ["cvss score","score","cvss"])),
            impact=get_col(row, ["impact","dampak","cvss 3.1 impact","cvss impact"]),
            mitigation=get_col(row, ["mitigation","mitigasi","remediation"]),
            status=get_col(row, ["status"]) or "Open",
            finding_date=parse_date(get_col(row, ["finding date","tanggal temuan","finding_date"])),
            resolution_date=parse_date(get_col(row, ["resolution date","resolution_date"])),
            fixing_date=parse_date(get_col(row, ["fixing date","fixing_date"])),
            referensi=get_col(row, ["referensi","reference","references"]),
            note=get_col(row, ["note","notes","catatan"]),
            created_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc),
        )
        db.add(r)
        imported += 1

    await db.commit()
    return {"imported": imported}

# ── Export Excel ──────────────────────────────────────────────
@router.get("/reports/export")
async def export_excel(
    company_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    result = await db.execute(select(VulnReport).where(VulnReport.company_id == company_id).order_by(VulnReport.no))
    reports = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"
    headers = ["Periode","Vuln Code","Vuln ID","Members","Vuln Name","Description","Severity",
               "CVSS Vector","CVSS Score","Impact","Mitigation","Status",
               "Finding Date","Resolution Date","Fixing Date","Referensi","Note"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="1e1e2e")
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(reports, 2):
        ws.append([
            r.no, r.vuln_code, r.vuln_id, r.vuln_members, r.vuln_name,
            r.description, r.severity, r.cvss_vector, r.cvss_score,
            r.impact, r.mitigation, r.status,
            r.finding_date.strftime("%Y-%m-%d") if r.finding_date else None,
            r.resolution_date.strftime("%Y-%m-%d") if r.resolution_date else None,
            r.fixing_date.strftime("%Y-%m-%d") if r.fixing_date else None,
            r.referensi, r.note,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=vulnerabilities_{company_id}.xlsx"})

# ── Stats ─────────────────────────────────────────────────────
@router.get("/stats")
async def get_stats(db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    companies_q = await db.execute(select(VulnCompany))
    companies = companies_q.scalars().all()
    result = []
    for c in companies:
        status_q = await db.execute(
            select(VulnReport.status, func.count(VulnReport.id))
            .where(VulnReport.company_id == c.id)
            .group_by(VulnReport.status)
        )
        sev_q = await db.execute(
            select(VulnReport.severity, func.count(VulnReport.id))
            .where(VulnReport.company_id == c.id)
            .group_by(VulnReport.severity)
        )
        result.append({
            "company": {"id": c.id, "name": c.name, "code": c.code, "color": c.color},
            "by_status": {row[0]: row[1] for row in status_q},
            "by_severity": {row[0]: row[1] for row in sev_q},
        })
    return result

def _report_dict(r: VulnReport) -> dict:
    return {
        "id": r.id, "company_id": r.company_id,
        "no": r.no, "vuln_code": r.vuln_code, "vuln_id": r.vuln_id,
        "vuln_members": r.vuln_members, "vuln_name": r.vuln_name,
        "description": r.description, "severity": r.severity,
        "cvss_vector": r.cvss_vector, "cvss_score": r.cvss_score,
        "cvss_version": r.cvss_version or "3.1",
        "impact": r.impact, "mitigation": r.mitigation, "status": r.status,
        "finding_date": r.finding_date.strftime("%Y-%m-%d") if r.finding_date else None,
        "resolution_date": r.resolution_date.strftime("%Y-%m-%d") if r.resolution_date else None,
        "fixing_date": r.fixing_date.strftime("%Y-%m-%d") if r.fixing_date else None,
        "referensi": r.referensi, "note": r.note,
        "product_id": r.product_id,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "updated_at": r.updated_at.isoformat() if r.updated_at else None,
    }

def _to_int(v):
    try: return int(v)
    except: return None

def _to_float(v):
    try: return float(v)
    except: return None

def _norm_sev(v):
    if not v: return "medium"
    v = v.lower()
    for s in ["critical","high","medium","low","informational","info"]:
        if s in v: return s
    return "medium"


# ── Products ──────────────────────────────────────────────────
from app.db.models import VulnProduct as _VulnProduct

class ProductCreate(BaseModel):
    company_id: str
    name: str

@router.get("/products")
async def get_products(
    company_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    q = select(_VulnProduct).order_by(_VulnProduct.name)
    if company_id:
        q = q.where(_VulnProduct.company_id == company_id)
    result = await db.execute(q)
    return [{"id": p.id, "company_id": p.company_id, "name": p.name} for p in result.scalars()]

@router.post("/products")
async def create_product(
    data: ProductCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    p = _VulnProduct(
        id=str(uuid.uuid4()),
        company_id=data.company_id,
        name=data.name,
        created_at=datetime.now(timezone.utc)
    )
    db.add(p)
    await db.commit()
    return {"id": p.id, "company_id": p.company_id, "name": p.name}

@router.delete("/products/{product_id}")
async def delete_product(
    product_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Clear product_id from reports
    from sqlalchemy import update as _update
    await db.execute(_update(VulnReport).where(VulnReport.product_id == product_id).values(product_id=None))
    await db.execute(delete(_VulnProduct).where(_VulnProduct.id == product_id))
    await db.commit()
    return {"ok": True}


@router.get("/product-stats")
async def get_product_stats(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.db.models import VulnProduct as _VP
    companies_q = await db.execute(select(VulnCompany))
    companies = companies_q.scalars().all()
    result = []
    for c in companies:
        prods_q = await db.execute(select(_VP).where(_VP.company_id == c.id))
        prods = prods_q.scalars().all()
        prod_data = []
        for p in prods:
            cnt_q = await db.execute(
                select(func.count(VulnReport.id))
                .where(VulnReport.product_id == p.id)
            )
            cnt = cnt_q.scalar() or 0
            if cnt > 0:
                prod_data.append({"name": p.name, "count": cnt, "product_id": p.id})
        result.append({"company_id": c.id, "products": prod_data})
    return result
